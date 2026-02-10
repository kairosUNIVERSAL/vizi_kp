import openpyxl
import json

file_path = r"c:\Assensio99\Code stuf\KP_LIGHT_VISO\price_list_ceiling.xlsx"

try:
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet_names = workbook.sheetnames
    
    result = {"sheets": {}}
    
    for sheet in sheet_names:
        ws = workbook[sheet]
        
        # Get first few rows
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 5: break
            rows.append(row)
            
        result["sheets"][sheet] = {
            "rows": rows
        }
        
    print(json.dumps(result, indent=2, default=str))

except Exception as e:
    print(f"Error: {e}")
