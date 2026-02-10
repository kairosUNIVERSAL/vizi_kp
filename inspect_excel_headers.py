import openpyxl
import os

EXCEL_PATH = r"c:\Assensio99\Code stuf\KP_LIGHT_VISO\price_list_ceiling.xlsx"

def inspect_headers():
    if not os.path.exists(EXCEL_PATH):
        print(f"File not found: {EXCEL_PATH}")
        return

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    print("Sheets found:", wb.sheetnames)
    for sheet in wb.sheetnames:
        print(f"\n--- Sheet: {sheet} ---")
        ws = wb[sheet]
        for i, row in enumerate(ws.iter_rows(max_row=5, values_only=True)):
             row_clean = [str(c).strip() for c in row if c]
             print(f"Row {i+1}: {row_clean}")

if __name__ == "__main__":
    inspect_headers()
