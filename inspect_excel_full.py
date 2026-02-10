import openpyxl
import os
import json

EXCEL_PATH = r"c:\Assensio99\Code stuf\KP_LIGHT_VISO\price_list_ceiling.xlsx"
OUTPUT_PATH = r"c:\Assensio99\Code stuf\KP_LIGHT_VISO\ceiling-kp\excel_analysis.json"

def inspect_excel():
    if not os.path.exists(EXCEL_PATH):
        print(f"File not found: {EXCEL_PATH}")
        return

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    
    result = {
        "file": EXCEL_PATH,
        "sheet_count": len(wb.sheetnames),
        "sheet_names": wb.sheetnames,
        "sheets": {}
    }
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_data = {
            "max_row": ws.max_row,
            "max_col": ws.max_column,
            "rows": []
        }
        
        # Read ALL rows (up to 200 to be safe)
        for i, row in enumerate(ws.iter_rows(max_row=min(ws.max_row, 200), values_only=True)):
            row_clean = []
            for cell in row:
                if cell is not None:
                    row_clean.append(str(cell).strip())
                else:
                    row_clean.append("")
            # Only add rows that have at least one non-empty cell
            if any(c for c in row_clean):
                sheet_data["rows"].append({
                    "row_num": i + 1,
                    "cells": row_clean
                })
        
        result["sheets"][sheet_name] = sheet_data
    
    # Write to JSON
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    
    print(f"Analysis saved to: {OUTPUT_PATH}")
    print(f"Sheets: {wb.sheetnames}")
    for sn in wb.sheetnames:
        sd = result["sheets"][sn]
        print(f"  {sn}: {sd['max_row']} rows x {sd['max_col']} cols, {len(sd['rows'])} non-empty rows")

if __name__ == "__main__":
    inspect_excel()
