"""
Script to import price items from Excel file.
Run this script on the server with: python scripts/import_prices.py

The script reads ceiling_price_research.xlsx and imports items into the database.
"""

import sys
import os

# Add project root to path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import load_workbook
from sqlalchemy.orm import Session
from app.database import SessionLocal
from app.models import PriceItem, Category

# Configuration
EXCEL_FILE = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'ceiling_price_research.xlsx')
DEFAULT_COMPANY_ID = 1  # Change this to your company ID

def get_or_create_category(db: Session, name: str) -> Category:
    """Get existing category or create new one."""
    slug = name.lower().replace(' ', '_').replace('-', '_')
    category = db.query(Category).filter(Category.slug == slug).first()
    if not category:
        max_order = db.query(Category).count()
        category = Category(
            name=name,
            slug=slug,
            sort_order=max_order + 1,
            is_system=False
        )
        db.add(category)
        db.commit()
        db.refresh(category)
        print(f"  Created category: {name}")
    return category


def import_prices(company_id: int = DEFAULT_COMPANY_ID):
    """Import prices from Excel file."""
    
    if not os.path.exists(EXCEL_FILE):
        print(f"Error: Excel file not found: {EXCEL_FILE}")
        return
    
    print(f"Loading Excel file: {EXCEL_FILE}")
    wb = load_workbook(EXCEL_FILE)
    sheet = wb.active
    
    # Get headers
    headers = [cell.value for cell in sheet[1]]
    print(f"Headers: {headers}")
    
    # Find column indices (adjust based on actual Excel structure)
    # Common columns: Название, Цена, Ед.изм., Категория, Синонимы
    name_col = None
    price_col = None
    unit_col = None
    category_col = None
    synonyms_col = None
    
    for idx, header in enumerate(headers):
        if header:
            h = header.lower().strip()
            if 'название' in h or 'наименование' in h or 'name' in h:
                name_col = idx
            elif 'цена' in h or 'price' in h or 'стоимость' in h:
                price_col = idx
            elif 'ед' in h or 'unit' in h:
                unit_col = idx
            elif 'категория' in h or 'category' in h or 'тип' in h:
                category_col = idx
            elif 'синоним' in h or 'synonym' in h or 'альтернатив' in h:
                synonyms_col = idx
    
    print(f"Column mapping: name={name_col}, price={price_col}, unit={unit_col}, category={category_col}, synonyms={synonyms_col}")
    
    if name_col is None:
        print("Error: Could not find 'Название' column")
        return
    
    db = SessionLocal()
    imported = 0
    skipped = 0
    
    try:
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                name = row[name_col] if name_col is not None else None
                if not name or str(name).strip() == '':
                    continue
                
                name = str(name).strip()
                
                # Check if already exists
                existing = db.query(PriceItem).filter(
                    PriceItem.company_id == company_id,
                    PriceItem.name == name
                ).first()
                
                if existing:
                    print(f"  Skipped (exists): {name}")
                    skipped += 1
                    continue
                
                # Get values
                price = 0
                if price_col is not None and row[price_col]:
                    try:
                        price = float(str(row[price_col]).replace(',', '.').replace(' ', '').replace('₽', ''))
                    except ValueError:
                        price = 0
                
                unit = 'шт'
                if unit_col is not None and row[unit_col]:
                    unit = str(row[unit_col]).strip()
                
                synonyms = ''
                if synonyms_col is not None and row[synonyms_col]:
                    synonyms = str(row[synonyms_col]).strip()
                
                # Get or create category
                category_name = 'Общее'
                if category_col is not None and row[category_col]:
                    category_name = str(row[category_col]).strip()
                
                category = get_or_create_category(db, category_name)
                
                # Create item
                item = PriceItem(
                    company_id=company_id,
                    category_id=category.id,
                    name=name,
                    price=price,
                    unit=unit,
                    synonyms=synonyms,
                    is_active=True,
                    is_custom=True
                )
                db.add(item)
                db.commit()
                
                print(f"  Imported: {name} ({price} ₽/{unit})")
                imported += 1
                
            except Exception as e:
                print(f"  Error on row {row_idx}: {e}")
                db.rollback()
    
    finally:
        db.close()
    
    print(f"\n=== Import Complete ===")
    print(f"Imported: {imported}")
    print(f"Skipped: {skipped}")


if __name__ == '__main__':
    # Allow passing company_id as argument
    company_id = DEFAULT_COMPANY_ID
    if len(sys.argv) > 1:
        try:
            company_id = int(sys.argv[1])
        except ValueError:
            pass
    
    print(f"Importing prices for company_id={company_id}")
    import_prices(company_id)
