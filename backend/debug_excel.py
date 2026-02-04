import openpyxl
import json
import os

file_path = "ceiling_price_research.xlsx"
output_path = "debug_excel_output.json"

if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    exit(1)

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    result = {"sheets": {}}
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        data = []
        # Read header
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
            
        header = rows[0]
        data.append(list(header))
        
        # Read first 20 rows of data
        for row in rows[1:30]:
            # Filter out completely empty rows
            if any(cell is not None for cell in row):
                data.append(list(row))
                
        result["sheets"][sheet_name] = data
        
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2, default=str)
        
    print(f"Successfully wrote to {output_path}")

except Exception as e:
    print(f"Error reading excel: {e}")
