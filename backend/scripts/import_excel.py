"""
Import script for price_list_ceiling.xlsx

Excel Structure (4 sheets):
- Полотна:      № | Наименование | Ед. | Цена | Синонимы        -> category: "Полотна" (is_equipment=False)
- Профили:      № | Наименование | Ед. | Цена | Синонимы        -> category: "Профили" (is_equipment=False)
- Услуги:       № | Категория | Наименование | Ед. | Цена | Синонимы  -> category per "Категория" col (is_equipment=False)
- Оборудование: № | Категория | Наименование | Ед. | Цена | Синонимы | Комментарий -> category per "Категория" col (is_equipment=True)
"""
import sys
import os
import openpyxl
import logging

# Add backend directory to path so we can import app modules
sys.path.append(os.path.join(os.path.dirname(__file__), '..'))

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("import_log.txt", mode='w', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

from app.database import SessionLocal
from app.models import PriceItem, Category, Company, User

# Path to the Excel file (same directory as this script)
EXCEL_PATH = os.path.join(os.path.dirname(__file__), 'price_list_ceiling.xlsx')


def slugify(text: str) -> str:
    """Create a simple slug from Russian text."""
    mapping = {
        'а': 'a', 'б': 'b', 'в': 'v', 'г': 'g', 'д': 'd', 'е': 'e', 'ё': 'yo',
        'ж': 'zh', 'з': 'z', 'и': 'i', 'й': 'y', 'к': 'k', 'л': 'l', 'м': 'm',
        'н': 'n', 'о': 'o', 'п': 'p', 'р': 'r', 'с': 's', 'т': 't', 'у': 'u',
        'ф': 'f', 'х': 'kh', 'ц': 'ts', 'ч': 'ch', 'ш': 'sh', 'щ': 'shch',
        'ъ': '', 'ы': 'y', 'ь': '', 'э': 'e', 'ю': 'yu', 'я': 'ya',
        ' ': '_', '-': '_', '.': '', '/': '_',
    }
    result = []
    for char in text.lower():
        result.append(mapping.get(char, char))
    slug = ''.join(result)
    # Remove non-alphanumeric (keep underscores)
    slug = ''.join(c for c in slug if c.isalnum() or c == '_')
    # Collapse multiple underscores
    while '__' in slug:
        slug = slug.replace('__', '_')
    return slug.strip('_')


def get_or_create_category(db, name: str, is_equipment: bool) -> Category:
    """Find or create a category by name."""
    slug = slugify(name)
    category = db.query(Category).filter(Category.slug == slug).first()
    
    if not category:
        # Also try by name
        category = db.query(Category).filter(Category.name == name).first()
    
    if category:
        # Update flags
        category.is_equipment = is_equipment
        if category.slug != slug:
            category.slug = slug
        db.add(category)
        db.commit()
        logger.info(f"  Updated category: {name} (slug={slug}, is_equipment={is_equipment})")
    else:
        category = Category(
            name=name,
            slug=slug,
            is_system=True,
            is_equipment=is_equipment
        )
        db.add(category)
        db.commit()
        db.refresh(category)
        logger.info(f"  Created category: {name} (slug={slug}, is_equipment={is_equipment})")
    
    return category


def upsert_price_item(db, company_id: int, category_id: int, name: str, unit: str, price: float, synonyms: str):
    """Create or update a price item."""
    item = db.query(PriceItem).filter(
        PriceItem.company_id == company_id,
        PriceItem.category_id == category_id,
        PriceItem.name == name
    ).first()

    if item:
        item.price = price
        item.unit = unit
        item.synonyms = synonyms
        item.is_active = True
    else:
        item = PriceItem(
            company_id=company_id,
            category_id=category_id,
            name=name,
            unit=unit,
            price=price,
            synonyms=synonyms,
            is_active=True,
            is_custom=False
        )
        db.add(item)


def parse_price(value) -> float:
    """Safely parse a price value."""
    if value is None:
        return 0.0
    try:
        return float(str(value).replace(" ", "").replace(",", "."))
    except (ValueError, TypeError):
        return 0.0


def import_simple_sheet(db, ws, company_id: int, category, sheet_name: str):
    """
    Import a simple sheet (Полотна, Профили).
    Columns: № | Наименование | Ед. | Цена | Синонимы
    """
    count = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        # row[0]=№, row[1]=Наименование, row[2]=Ед., row[3]=Цена, row[4]=Синонимы
        if not row[1]:
            continue

        name = str(row[1]).strip()
        if not name:
            continue

        unit = str(row[2]).strip() if row[2] else "шт"
        price = parse_price(row[3])
        synonyms = str(row[4]).strip() if row[4] else ""

        # Add sheet name to synonyms for searchability
        if sheet_name and sheet_name.lower() not in synonyms.lower():
            synonyms = f"{sheet_name}, {synonyms}" if synonyms else sheet_name

        upsert_price_item(db, company_id, category.id, name, unit, price, synonyms)
        count += 1

    db.commit()
    logger.info(f"  Sheet '{sheet_name}' → category '{category.name}': imported {count} items")


def import_categorized_sheet(db, ws, company_id: int, category, sheet_name: str):
    """
    Import a sheet with subcategories (Услуги, Оборудование).
    All items go into the provided category. Subcategory column goes to synonyms.
    Columns: № | Категория | Наименование | Ед. | Цена | Синонимы [| Комментарий]
    """
    count = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        # row[0]=№, row[1]=Категория (subcategory), row[2]=Наименование, row[3]=Ед., row[4]=Цена, row[5]=Синонимы
        if not row[2]:
            continue

        subcategory = str(row[1]).strip() if row[1] else ""
        name = str(row[2]).strip()
        if not name:
            continue

        unit = str(row[3]).strip() if row[3] else "шт"
        price = parse_price(row[4])
        synonyms = str(row[5]).strip() if row[5] else ""

        # Append subcategory to synonyms for search (e.g. "Углы", "Карнизы")
        if subcategory and subcategory.lower() not in synonyms.lower():
            synonyms = f"{subcategory}, {synonyms}" if synonyms else subcategory

        upsert_price_item(db, company_id, category.id, name, unit, price, synonyms)
        count += 1

    db.commit()
    logger.info(f"  Sheet '{sheet_name}' → category '{category.name}': imported {count} items")


def import_price_list():
    """Main import function."""
    db = SessionLocal()
    try:
        if not os.path.exists(EXCEL_PATH):
            logger.error(f"File not found: {EXCEL_PATH}")
            return

        logger.info(f"Loading Excel: {EXCEL_PATH}")
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        logger.info(f"Sheets found: {wb.sheetnames}")

        # Find company
        user = db.query(User).first()
        if not user or not user.company:
            logger.error("No user/company found in database. Cannot import.")
            return

        company_id = user.company.id
        logger.info(f"Importing for company: {user.company.name} (ID: {company_id})")

        # ── Cleanup: deactivate ALL old items (safe for FK references in estimates) ──
        deactivated = db.query(PriceItem).filter(
            PriceItem.company_id == company_id
        ).update({"is_active": False})
        db.commit()
        logger.info(f"Deactivated {deactivated} old price items")

        # Delete categories with no linked estimate items (orphans from old seed)
        from sqlalchemy import func, and_
        from app.models import EstimateItem
        orphan_cats = (
            db.query(Category)
            .outerjoin(PriceItem, PriceItem.category_id == Category.id)
            .outerjoin(EstimateItem, EstimateItem.price_item_id == PriceItem.id)
            .group_by(Category.id)
            .having(func.count(EstimateItem.id) == 0)
            .all()
        )
        for cat in orphan_cats:
            # Delete price items in this category first
            db.query(PriceItem).filter(PriceItem.category_id == cat.id).delete()
            db.delete(cat)
        db.commit()
        if orphan_cats:
            logger.info(f"Deleted {len(orphan_cats)} orphaned categories: {[c.name for c in orphan_cats]}")

        # ── Business rule: 2 categories ──
        # Услуги (is_equipment=false) = sheet "Услуги"
        # Оборудование (is_equipment=true) = everything else (Полотна, Профили, Оборудование, etc.)
        cat_services = get_or_create_category(db, "Услуги", is_equipment=False)
        cat_equipment = get_or_create_category(db, "Оборудование", is_equipment=True)

        # Process each sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            logger.info(f"Processing sheet: {sheet_name} ({ws.max_row} rows)")

            # Read header row to determine column layout
            header = [str(c).strip().lower() if c else "" for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
            has_category_col = "категория" in header

            # Determine target category
            if sheet_name.lower() == "услуги":
                target_category = cat_services
            else:
                target_category = cat_equipment

            if has_category_col:
                # Sheets with subcategory column (Услуги, Оборудование)
                import_categorized_sheet(db, ws, company_id, target_category, sheet_name)
            else:
                # Simple sheets (Полотна, Профили) — sheet name goes to synonyms
                import_simple_sheet(db, ws, company_id, target_category, sheet_name)

        logger.info("Import completed successfully!")

    except Exception as e:
        logger.exception(f"Error importing excel: {e}")
    finally:
        db.close()


if __name__ == "__main__":
    import_price_list()
