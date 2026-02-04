import openpyxl
import sys
import os

# Create app directory context
sys.path.append(os.getcwd())

from sqlalchemy.orm import Session
from app.database import SessionLocal, engine, Base
from app.models import PriceItem, Category, Company, User
from app.api.deps import get_db

def seed_data():
    log_file = open("seed_result.txt", "w", encoding="utf-8")
    
    def log(msg):
        print(msg)
        log_file.write(msg + "\n")
        log_file.flush()

    log("Starting seed_db.py...")
    db = SessionLocal()
    try:
        user = db.query(User).filter(User.email == "admin@example.com").first()
        if not user:
            # Try getting any user
            user = db.query(User).first()
            if not user:
                log("No users found. Creating admin...")
                from app.services.auth_service import get_password_hash
                user = User(
                    email="admin@example.com", 
                    hashed_password=get_password_hash("admin"), 
                    full_name="Admin",
                    is_active=True
                )
                db.add(user)
                db.commit()
                db.refresh(user)
        
        log(f"Using user: {user.email}")
            
        if not user.company:
            log("User has no company. Creating...")
            company = Company(user_id=user.id, name="Potolok Service", city="Moscow")
            db.add(company)
            db.commit()
            db.refresh(company)
        else:
            company = user.company
            
        company_id = company.id
        log(f"Seeding for company: {company.name} (ID: {company_id})")

        file_path = "ceiling_price_research.xlsx"
        if not os.path.exists(file_path):
            log(f"ERROR: File not found at {os.path.abspath(file_path)}")
            return

        wb = openpyxl.load_workbook(file_path, data_only=True)
        log(f"Workbook loaded. Sheets: {wb.sheetnames}")

        sheet_mappings = [
            ("Полотна", "Полотна (м²)", "canvases", 10),
            ("Услуги", "Работы и профили", "services", 20),
            ("Освещение", "Освещение", "lighting", 30),
            ("Профили", "Комплектующие", "components", 40)
        ]

        total_added = 0
        total_updated = 0

        for sheet_name, cat_name, cat_slug, sort_order in sheet_mappings:
            if sheet_name not in wb.sheetnames:
                log(f"Skipping {sheet_name} - not found")
                continue
            
            # Find category by slug first
            category = db.query(Category).filter(Category.slug == cat_slug).first()
            if not category:
                category = db.query(Category).filter(Category.name == cat_name).first()
                if category:
                    # found by name, ensure slug
                    if category.slug != cat_slug:
                        category.slug = cat_slug
                        db.commit()
                else:
                    # Create new
                    category = Category(name=cat_name, slug=cat_slug, sort_order=sort_order)
                    db.add(category)
                    db.commit()
                    db.refresh(category)
            
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            log(f"Processing {sheet_name}: Found {len(rows)} rows")
            
            for i, row in enumerate(rows):
                if not row or not row[1]: 
                    continue

                item_data = {}
                try:
                    if sheet_name == "Полотна":
                        brand = str(row[1] or '')
                        ftype = str(row[2] or '')
                        width = str(row[3] or '')
                        texture = str(row[4] or '')
                        color = str(row[5] or '')
                        price = float(row[6] or 0)
                        
                        full_name = f"{brand} {texture} {color} {width}"
                        # Append other details to synonyms for search
                        extra_synonyms = f"{brand}, {texture}, {color}, {ftype}"
                        
                        item_data = {
                            "name": full_name.strip(),
                            "price": price,
                            "unit": "м²",
                            "synonyms": extra_synonyms
                        }
                    elif sheet_name == "Услуги":
                        name = str(row[2] or '')
                        unit = str(row[3] or 'шт')
                        price = float(row[4] or 0)
                        synonyms = str(row[5] or '')
                        if synonyms == 'None': synonyms = ""
                        item_data = {"name": name, "price": price, "unit": unit, "synonyms": synonyms}
                    elif sheet_name == "Освещение":
                        name = str(row[2] or '')
                        price = float(row[3] or 0)
                        note = str(row[4] or '')
                        if note == 'None': note = ""
                        item_data = {"name": name, "price": price, "unit": "шт", "synonyms": note}
                    elif sheet_name == "Профили":
                        name = str(row[2] or '')
                        unit = str(row[3] or 'шт')
                        price = float(row[4] or 0)
                        note = str(row[5] or '')
                        if note == 'None': note = ""
                        item_data = {"name": name, "price": price, "unit": unit, "synonyms": note}

                    if item_data.get("name"):
                        existing = db.query(PriceItem).filter(
                            PriceItem.company_id == company_id,
                            PriceItem.name == item_data["name"]
                        ).first()

                        if existing:
                            existing.price = item_data["price"]
                            existing.unit = item_data["unit"]
                            if item_data["synonyms"]: existing.synonyms = item_data["synonyms"]
                            # No description field update
                            existing.category_id = category.id
                            existing.is_active = True
                            total_updated += 1
                        else:
                            new_item = PriceItem(
                                company_id=company_id,
                                category_id=category.id,
                                name=item_data["name"],
                                price=item_data["price"],
                                unit=item_data["unit"],
                                synonyms=item_data.get("synonyms", ""),
                                # No description field
                                is_active=True,
                                is_custom=False
                            )
                            db.add(new_item)
                            total_added += 1
                            
                except Exception as e:
                    log(f"Error parsing row {i} in {sheet_name}: {e}")

            db.commit()
            log(f"Finished {sheet_name}. Added: {total_added}, Updated: {total_updated}")

    except Exception as e:
        log(f"CRITICAL ERROR: {e}")
        import traceback
        traceback.print_exc(file=log_file)
    finally:
        db.close()
        log("Seed finished.")
        log_file.close()

if __name__ == "__main__":
    seed_data()
